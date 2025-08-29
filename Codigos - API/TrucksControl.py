import requests
import zipfile
import io
import os
import time
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xml.etree.ElementTree as eT
import pandas as pd

ultima_requisicoes = {}

def requisicao_xml(str_request, requisicao, podeRequisitar):
    url = "https://webservice.newrastreamentoonline.com.br"
    
    try:
        headers = {
            "Content-Type" : "text/xml"
        }
        
        resposta = requests.post(url, data=str_request, headers=headers)
        
        if resposta.status_code == 200:
            try:
                with io.BytesIO(resposta.content) as dados:
                    with zipfile.ZipFile(dados) as zip_file:
                        nome_arquivos = zip_file.namelist()

                        nome_txt = None
                        for name in nome_arquivos:
                            if name.endswith('.txt'):
                                nome_txt = name
                                break

                        if nome_txt:
                            print("entrou")
                            if podeRequisitar:
                                with zip_file.open(nome_txt) as arquivo:
                                    criar_atualizar_tabelas(arquivo.read(), requisicao)
                            else:
                                print("Não pode requisitar ainda")

            except zipfile.BadZipFile:
                print("Não é arquivo zip")
                criar_atualizar_tabelas(resposta.content, requisicao)
        else:
            print(f"Erro: {resposta.status_code}")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def criar_atualizar_tabelas(arquivo, requisicao):
    dados = []
    arquivo = arquivo.strip()
    root = eT.fromstring(arquivo)
    print(arquivo)

    for tags in root:
        item = {}
        for tag_menor in tags:
            item[tag_menor.tag] = tag_menor.text
        dados.append(item)
    
    df = pd.DataFrame(dados)
    df_limpo = df.drop_duplicates()
    df_final = renomear_colunas(df_limpo, requisicao)

    caminho_arquivo = f"dados_{requisicao}.xlsx"

    if os.path.exists(caminho_arquivo):
        book = load_workbook(caminho_arquivo)
        
        if 'Sheet1' in book.sheetnames:
            sheet = book['Sheet1']
            last_row = sheet.max_row
            for row in df_final.values.tolist():
                sheet.append(row)
        else:
            df_final.to_excel(caminho_arquivo, index=False)
    else:
        df_final.to_excel(caminho_arquivo, index=False)

    book.save(caminho_arquivo)

def renomear_colunas(planilha, requisicao):
    if requisicao == "RequestAlertasSoftware":
        planilha_atualizada = planilha.rename(columns={"alsID" : "ID do alerta", 
                                                       "mID" : "ID da mensagem", 
                                                       "veiID" : "ID do veiculo", 
                                                       "tdeID" : "ID do alerta de software", 
                                                       "dtInc" : "Data da entrada",
                                                       "dt" : "Data de saida", 
                                                       "desc" : "Cerca eletronica", 
                                                       "odo" : "Odometro"})
    if requisicao == "RequestDescricaoAlertas":
        planilha_atualizada = planilha.rename(columns={"tdeID" : "ID do alerta de software"})
    if requisicao == "RequestAcessorio":
        planilha_atualizada = planilha.rename(columns={"acvID" : "ID do acessorio"})
    if requisicao == "RequestAcessorioVeiculo":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo",
                                                       "acvID" : "ID do acessorio"})
    if requisicao == "RequestVeiculoEspelhado":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo", 
                                                       "cmd" : "Premissão para o envio de comando",
                                                       "IE" : "Possui inteligencia embarcada", 
                                                       "TIE" : "Transferencia de inteligencia embarcada",
                                                       "cgccpf" : "CNPJ/CPF do cliente", 
                                                       "possocancelar" : "Pode cancelar?"})
    if requisicao == "RequestGrupoMacro":
        planilha_atualizada = planilha.rename(columns={"tgrID" : "ID do grupo",
                                                       "nm" : "Nome do grupo",
                                                       "dGrupo" : "Descrição do grupo", 
                                                       "dt" : "Data de criação do grupo",
                                                       "emb" : "Pode embacar nesse grupo? (0 = N, 1 = S)",
                                                       "tec" : "Qual teclado o grupo foi criado? (1 = Teclado Grande; 2 = Teclado Pequeno)"})
    if requisicao == "RequestItemMacro":
        planilha_atualizada = planilha.rename(columns={"tfrID" : "ID da macro",
                                                       "tgrID" : "ID do grupo",
                                                       "nm" : "Nome da macro",
                                                       "cod" : "Codigo da macro no CB",
                                                       "cont" : "Conteudo da macro",
                                                       "exc" : "Status (0 = Ativo, 1 = Excluído)",
                                                       "aut" : "Autenticação (0 = Não exige, 1 = Exige)"})
    if requisicao == "RequestGrupoMacroEmbarcado":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo",
                                                       "ativoCV" : "Ativo Central-Veiculo",
                                                       "ativoVC" : "Ativo Veiculo-Central",
                                                       "vc1" : "Grupo veiculo-central embarcado 1 (-1 = Vazio)",
                                                       "vc2" : "Grupo veiculo-central embarcado 2 (-1 = Vazio)",
                                                       "vc3" : "Grupo veiculo-central embarcado 3 (-1 = Vazio)"})
    if requisicao == "RequestVeiculo":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo",
                                                       "placa" : "Placa do veiculo",
                                                       "vs" : "Versao do computador",
                                                       "tCmd" : "Tempo para reenvio de comando",
                                                       "tMac" : "Possui teclado de macro",
                                                       "eCmd" : "Tem permissão para enviar comando?",
                                                       "tp" : "Temporizador padrao",
                                                       "ta" : "Temporizador satelital",
                                                       "eqp" : "Tipo do equipamento",
                                                       "prop" : "O cliente e proprietario?",
                                                       "dIE" : "O cliente tem direito a inteligência?",
                                                       "loc" : "Equipamento esta sinalizando?",
                                                       "ident" : "Prefixo",
                                                       "vManut" : "Esta em manutenção? (0 = N, 1 = S)",
                                                       "podeCompartilhar" : "Pode compartilhar info.",
                                                       "tgsm" : "Temporizador Gsm contratado",
                                                       "ppc" : "Permitido alterar posição satelital?",
                                                       "tppc" : "Temporizador posicionamento parado (-1 = N, 0 = S)",
                                                       "ppcMenor60" : "Temporizador posicionamento parado menor que 60 minutos (-1 = N, 0 = S)",
                                                       "IE" : "Cliente está com a inteligência?",
                                                       "mot" : "Nome do motorista"})
    if requisicao == "RequestVeiculoRedundante":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo",
                                                       "veiID2" : "ID do equipamento redundancia"})
    if requisicao == "RequestProjetoPontoControle":
        planilha_atualizada = planilha.rename(columns={"prjID" : "ID do projeto de ponto de controle",
                                                       "prjNome" : "Nome do projeto do ponto de controle"})
    if requisicao == "RequestCercaEletronica":
        planilha_atualizada = planilha.rename(columns={"cerID" : "ID da cerca eletrônica",
                                                       "nm" : "Nome da cerca eletrônica",
                                                       "des" : "Descrição",
                                                       "dt" : "Data da criação da cerca",
                                                       "dtmod" : "Data de modificação da cerca"})
    if requisicao == "RequestTelemetriaProjetos":
        planilha_atualizada = planilha.rename(columns={"tlpID" : "ID do projeto de telemetria",
                                                       "dtCriacao" : "Data de criação",
                                                       "dtModif" : "Data de modificação"})
    if requisicao == "RequestTelemetriaItem":
        planilha_atualizada = planilha.rename(columns={"tteID" : "ID do item da telemetria",
                                                       "tteDescricao" : "Descrição"})
    if requisicao == "RequestMotorista":
        planilha_atualizada = planilha.rename(columns={"motID" : "ID do motorista",
                                                       "mot" : "Nome do motorista",
                                                       "rg" : "RG",
                                                       "cpf" : "CPF"})
    if requisicao == "RequestMotoristaEmbarcado":
        planilha_atualizada = planilha.rename(columns={"veiID" : "ID do veiculo",
                                                       "motID" : "ID do motorista"})
    else:
        return planilha
    
    return planilha_atualizada

def requisicoes():
    todas_requisicoes = {
        "RequestAlertasSoftware" : 60,
        "RequestDescricaoAlertas" : 60,
        "RequestAcessorio" : 5 * 60,
        "RequestAcessorioVeiculo" : 5 * 60,
        "RequestSpyEspelhado" : 60,
        "RequestEspelhamentoPendenteSpy" : 60,
        "RequestVeiculoEspelhado" : 5 * 60,
        "RequestEspelhamentoPendenteVeiculo" : 60,
        "RequestGrupoMacro" : 30,
        "RequestItemMacro" : 5 * 60,
        "RequestGrupoMacroEmbarcado" : 5 * 60,
        "RequestSpy" : 5 * 60,
        "RequestVeiculo" : 5 * 60,
        "RequestVeiculoRedundante" : 5 * 60,
        "RequestMensagemCB" : 30,
        "RequestReferenciaEntrega" : 10,
        "RequestStatuscmie" : 30,
        "RequestPontoControle" : 30,
        "RequestProjetoPontoControle" : 60,
        "RequestCercaEletronica" : 5 * 60,
        "RequestCercaEletronicaEmbarcada" : 5 * 60,
        "RequestTelemetriaProjetos" : 60,
        "RequestTelemetriaItem" : 24 * 60 * 60,
        "RequestDataHoraServidor" : 60,
        "RequestMotorista" : 60,
        "RequestMotoristaEmbarcado" : 60
    }

    for requisicao, intervalo in todas_requisicoes.items():
        ultima_requisicao = ultima_requisicoes.get(requisicao)
        podeRequisitar = True
        if ultima_requisicao:
            print("passou pela ultima_requisicao")
            tempo_decorrido = datetime.now() - ultima_requisicao
            if tempo_decorrido.total_seconds() < intervalo:
                tempo_restante = intervalo - tempo_decorrido.total_seconds()
                podeRequisitar = False
                print(f"Ainda está no intervalo da requisição: {requisicao}, falta ainda: {tempo_restante:.2f}")
                continue
            else:
                podeRequisitar = True
                print(f"Realizando a requisição: {requisicao}")
        else:
            print(f"Primeira vez realizando a requisição: {requisicao}")

        if requisicao == "RequestAcessorioVeiculo" or requisicao == "RequestItemMacro":
            xml_request = f"""<?xml version="1.0" encoding="UTF-8"?>
                                    <{requisicao}>
                                        <login>Confidencial</login>
                                        <senha>Confidencial</senha>
                                        <todosItens>1</todosItens>
                                    </{requisicao}>"""
        if requisicao == "RequestMensagemCB":
            xml_request = f"""<?xml version="1.0" encoding="UTF-8"?>
                        <{requisicao}>
                            <login>Confidencial</login>
                            <senha>Confidencial</senha>
                            <mId>1</mId>
                        </{requisicao}>"""
        else:
            xml_request = f"""<?xml version="1.0" encoding="UTF-8"?>
                                <{requisicao}>
                                    <login>Confidencial</login>
                                    <senha>Confidencial</senha>
                                </{requisicao}>"""

        requisicao_xml(xml_request, requisicao, podeRequisitar)

        ultima_requisicoes[requisicao] = datetime.now()
        
        print(f"Requisição para {requisicao} finalizada. Próxima requisição será realizada quando o tempo de espera passar.")

while True:
    requisicoes()
    time.sleep(31)